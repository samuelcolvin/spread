#!/usr/bin/env -S uv run
# /// script
# dependencies = [
#   "openpyxl>=3.1,<4",
# ]
# ///

from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_CELL_BG = "ffffff"
DEFAULT_CELL_TEXT = "202124"
DEFAULT_COLUMN_WIDTH = 120.0
DEFAULT_ROW_HEIGHT = 24.0


@dataclass
class InspectedCell:
    x: str
    y: int
    display_value: str
    fg: str
    bg: str
    bold: bool
    width: float
    height: float


@dataclass
class InspectedSheet:
    sheet: str
    rows: int
    cols: int
    cells: list[InspectedCell]


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect an XLSX sheet display model.")
    parser.add_argument("path", type=Path)
    parser.add_argument("--sheet", help="Sheet name; defaults to the active sheet")
    parser.add_argument("--max-rows", type=int)
    parser.add_argument("--max-cols", type=int)

    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--json", action="store_true", help="Print JSON for strict comparison")
    mode.add_argument("--terminal", action="store_true", help="Print compact debug rows")

    args = parser.parse_args()
    sheet = inspect_xlsx(args.path, args.sheet, args.max_rows, args.max_cols)

    if args.json:
        print(json.dumps(asdict(sheet), indent=2, sort_keys=True))
    else:
        for cell in sheet.cells:
            print(
                "Cell{"
                f"x={cell.x}, y={cell.y}, fg={cell.fg}, bg={cell.bg}, "
                f"bold={str(cell.bold).lower()}, width={cell.width}, height={cell.height}, "
                f"display_value={cell.display_value!r}"
                "}"
            )


def inspect_xlsx(
    path: Path, sheet_name: str | None, max_rows: int | None, max_cols: int | None
) -> InspectedSheet:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    row_limit = min(worksheet.max_row, max_rows or worksheet.max_row)
    col_limit = min(worksheet.max_column, max_cols or worksheet.max_column)
    display_values = [
        [
            display_value(
                worksheet.cell(row=row_ix, column=col_ix).value,
                worksheet.cell(row=row_ix, column=col_ix).number_format,
            )
            for col_ix in range(1, col_limit + 1)
        ]
        for row_ix in range(1, row_limit + 1)
    ]

    rows = row_limit if max_rows else last_non_empty_row(display_values)
    cols = col_limit if max_cols else last_non_empty_col(display_values[:rows])

    cells: list[InspectedCell] = []
    for row_ix in range(1, rows + 1):
        row_height = row_height_px(worksheet, row_ix)
        for col_ix in range(1, cols + 1):
            column = get_column_letter(col_ix)
            cell = worksheet.cell(row=row_ix, column=col_ix)
            cells.append(
                InspectedCell(
                    x=column,
                    y=row_ix,
                    display_value=display_values[row_ix - 1][col_ix - 1],
                    fg=font_color(cell),
                    bg=fill_color(cell),
                    bold=bool(cell.font.bold),
                    width=column_width_px(worksheet, column),
                    height=row_height,
                )
            )

    return InspectedSheet(
        sheet=worksheet.title,
        rows=rows,
        cols=cols,
        cells=cells,
    )


def last_non_empty_row(display_values: list[list[str]]) -> int:
    for row_ix in range(len(display_values), 0, -1):
        if any(display_values[row_ix - 1]):
            return row_ix
    return 0


def last_non_empty_col(display_values: list[list[str]]) -> int:
    if not display_values:
        return 0
    for col_ix in range(len(display_values[0]), 0, -1):
        if any(row[col_ix - 1] for row in display_values):
            return col_ix
    return 0


def display_value(value: Any, number_format: str) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if (
            value.hour == 0
            and value.minute == 0
            and value.second == 0
            and value.microsecond == 0
        ):
            return value.date().isoformat()
        return value.isoformat(timespec="milliseconds").rstrip("0").rstrip(".")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, (int, float)):
        if is_percentage_format(number_format):
            return format_percentage(float(value), format_decimals(number_format))
        if is_dollar_format(number_format):
            return format_currency(float(value), format_decimals(number_format))
        if float(value).is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def is_percentage_format(number_format: str) -> bool:
    return "%" in number_format


def is_dollar_format(number_format: str) -> bool:
    lowered = number_format.lower()
    return "$" in lowered or "[$$" in lowered


def format_decimals(number_format: str) -> int:
    first_section = number_format.split(";", 1)[0]
    if "." not in first_section:
        return 0
    after_dot = first_section.split(".", 1)[1]
    count = 0
    for ch in after_dot:
        if ch in "0#":
            count += 1
        else:
            break
    return count


def format_percentage(value: float, decimals: int) -> str:
    return f"{value * 100:.{decimals}f}%"


def format_currency(value: float, decimals: int) -> str:
    sign = "-" if value < 0 else ""
    formatted = f"{abs(value):,.{decimals}f}"
    return f"{sign}${formatted}"


def font_color(cell: Any) -> str:
    if not cell.has_style:
        return DEFAULT_CELL_TEXT
    return rgb_color(cell.font.color) or DEFAULT_CELL_TEXT


def fill_color(cell: Any) -> str:
    if not cell.has_style:
        return DEFAULT_CELL_BG
    if cell.fill.fill_type != "solid":
        return DEFAULT_CELL_BG
    return rgb_color(cell.fill.fgColor) or DEFAULT_CELL_BG


def rgb_color(color: Any) -> str | None:
    if color is None or color.type != "rgb" or color.rgb is None:
        return None
    value = color.rgb
    if len(value) == 8:
        if value[:2] != "FF":
            return None
        value = value[2:]
    return value.lower()


def column_width_px(worksheet: Any, column: str) -> float:
    dimension = worksheet.column_dimensions.get(column)
    width = dimension.width if dimension is not None else None
    if width is None:
        width = worksheet.sheet_format.defaultColWidth
    if width is None:
        return DEFAULT_COLUMN_WIDTH
    return round(max(width * 7.0 + 5.0, 24.0), 2)


def row_height_px(worksheet: Any, row: int) -> float:
    height = worksheet.row_dimensions[row].height
    if height is None:
        height = worksheet.sheet_format.defaultRowHeight
    if height is None:
        return DEFAULT_ROW_HEIGHT
    return round(max(height * 4.0 / 3.0, 12.0), 2)


if __name__ == "__main__":
    main()
